"""
Excel Export for Face Attendance System
Client-side download functionality
"""

import pandas as pd
from datetime import date, timedelta
from typing import List, Optional
from sqlalchemy.orm import Session
from fastapi.responses import StreamingResponse
from openpyxl.styles import Font, PatternFill
import io

from app.models import Attendance, Student


def export_attendance_excel(
    db: Session,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    school_name: Optional[str] = None,
    class_name: Optional[str] = None,
    section: Optional[str] = None,
    student_ids: Optional[List[str]] = None
) -> StreamingResponse:
    """
    Export attendance data to Excel file for browser download
    """

    # Default today
    if not start_date:
        start_date = date.today()

    if not end_date:
        end_date = date.today()

    # Query
    query = db.query(Attendance, Student).join(
        Student,
        Attendance.student_id == Student.id
    )

    query = query.filter(
        Attendance.date >= start_date,
        Attendance.date <= end_date
    )

    if school_name:
        query = query.filter(Student.school_name == school_name)

    if class_name:
        query = query.filter(Student.class_name == class_name)

    if section:
        query = query.filter(Student.section == section)

    if student_ids:
        query = query.filter(Attendance.student_id.in_(student_ids))

    query = query.order_by(Attendance.date, Student.name)

    results = query.all()

    # Data build
    data = []

    if results:
        serial_no = 1

        for attendance, student in results:
            data.append({
                "Serial No": serial_no,
                "Name": student.name or "",
                "Class": student.class_name or "",
                "Section": student.section or "",
                "Enroll ID": student.id or "",
                "Roll": student.roll or "",
                "Date": attendance.date.strftime("%Y-%m-%d") if attendance.date else "",
                "In Time": attendance.in_time.strftime("%I:%M %p") if attendance.in_time else "",
                "Out Time": attendance.out_time.strftime("%I:%M %p") if attendance.out_time else "",
                "Status": attendance.status or "",
                "Remarks": attendance.remark or ""
            })
            serial_no += 1
    else:
        data.append({
            "Serial No": "",
            "Name": "",
            "Class": "",
            "Section": "",
            "Enroll ID": "",
            "Roll": "",
            "Date": "",
            "In Time": "",
            "Out Time": "",
            "Status": "",
            "Remarks": ""
        })

    df = pd.DataFrame(data)

    # Excel output
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        sheet_name = "Attendance Report"

        df.to_excel(
            writer,
            index=False,
            sheet_name=sheet_name,
            startrow=1
        )

        worksheet = writer.sheets[sheet_name]

        # Title row
        worksheet["A1"] = f"Attendance Report - {start_date} to {end_date}"
        worksheet.merge_cells("A1:K1")

        worksheet["A1"].font = Font(
            bold=True,
            size=14
        )

        # Header style
        fill = PatternFill(
            fill_type="solid",
            start_color="DDDDDD",
            end_color="DDDDDD"
        )

        for cell in worksheet[2]:
            cell.font = Font(bold=True)
            cell.fill = fill

        # Auto width
        for column in worksheet.columns:
            max_length = 0
            col_letter = column[0].column_letter

            for cell in column:
                try:
                    value = str(cell.value) if cell.value else ""
                    if len(value) > max_length:
                        max_length = len(value)
                except:
                    pass

            worksheet.column_dimensions[col_letter].width = min(max_length + 2, 50)

    output.seek(0)

    filename = f"attendance_{start_date.strftime('%Y-%m-%d')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )


def export_today_attendance_excel(db: Session) -> StreamingResponse:
    """
    Export today's attendance
    """

    today = date.today()

    return export_attendance_excel(
        db=db,
        start_date=today,
        end_date=today
    )


def generate_summary_excel(
    db: Session,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    school_name: Optional[str] = None,
    class_name: Optional[str] = None,
    section: Optional[str] = None
) -> StreamingResponse:
    """
    Generate attendance summary Excel
    """

    if not start_date:
        start_date = date.today()

    if not end_date:
        end_date = date.today()

    # Students query
    query = db.query(Student)

    if school_name:
        query = query.filter(Student.school_name == school_name)

    if class_name:
        query = query.filter(Student.class_name == class_name)

    if section:
        query = query.filter(Student.section == section)

    students = query.all()

    # Date range
    date_range = []
    current = start_date

    while current <= end_date:
        date_range.append(current)
        current += timedelta(days=1)

    # Summary data
    summary_data = []
    serial_no = 1

    for student in students:
        total_days = len(date_range)
        present_days = 0
        absent_days = 0
        leave_days = 0
        half_days = 0

        for check_date in date_range:
            attendance = db.query(Attendance).filter(
                Attendance.student_id == student.id,
                Attendance.date == check_date
            ).first()

            if attendance:
                if attendance.status == "P":
                    present_days += 1
                elif attendance.status == "A":
                    absent_days += 1
                elif attendance.status in ["H", "HD"]:
                    half_days += 1
                elif attendance.status == "L":
                    leave_days += 1

        attendance_percent = round(
            (present_days / total_days) * 100,
            2
        ) if total_days > 0 else 0

        summary_data.append({
            "Serial No": serial_no,
            "Enroll ID": student.id or "",
            "Student Name": student.name or "",
            "Class": student.class_name or "",
            "Section": student.section or "",
            "Roll Number": student.roll or "",
            "Total Days": total_days,
            "Present": present_days,
            "Absent": absent_days,
            "Half Day": half_days,
            "Leave": leave_days,
            "Attendance %": attendance_percent
        })

        serial_no += 1

    df = pd.DataFrame(summary_data)

    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        sheet_name = "Summary Report"

        df.to_excel(
            writer,
            index=False,
            sheet_name=sheet_name,
            startrow=1
        )

        worksheet = writer.sheets[sheet_name]

        # Correct title merge (IMPORTANT FIX)
        worksheet["A1"] = f"Attendance Summary - {start_date} to {end_date}"
        worksheet.merge_cells("A1:L1")

        worksheet["A1"].font = Font(
            bold=True,
            size=14
        )

        fill = PatternFill(
            fill_type="solid",
            start_color="DDDDDD",
            end_color="DDDDDD"
        )

        for cell in worksheet[2]:
            cell.font = Font(bold=True)
            cell.fill = fill

        # Auto width
        for column in worksheet.columns:
            max_length = 0
            col_letter = column[0].column_letter

            for cell in column:
                try:
                    value = str(cell.value) if cell.value else ""
                    if len(value) > max_length:
                        max_length = len(value)
                except:
                    pass

            worksheet.column_dimensions[col_letter].width = min(max_length + 2, 50)

    output.seek(0)

    filename = (
        f"attendance_summary_"
        f"{start_date.strftime('%Y-%m-%d')}_to_"
        f"{end_date.strftime('%Y-%m-%d')}.xlsx"
    )

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )
